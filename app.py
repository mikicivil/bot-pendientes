"""
Bot de WhatsApp para registrar pendientes de oficina.
Usa Twilio + Flask para recibir mensajes y guardarlos en Excel.

FORMATO DE MENSAJE:
  /nueva [área] / [actividad] / [fecha dd-mm] / [responsable]
  /nueva Santiago / solicitar ampliación de plazo / 15-03 / miguel

COMANDOS:
  /nueva  - Registrar nueva actividad
  /ver    - Ver pendientes de un área (ej: /ver Santiago)
  /hoy    - Ver todas las actividades de hoy
  /areas  - Ver lista de áreas disponibles
  /ayuda  - Ver instrucciones de uso
"""

from flask import Flask, request
from twilio.twiml.messaging_response import MessagingResponse
from datetime import datetime, date
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import re

app = Flask(__name__)

EXCEL_FILE = os.environ.get("EXCEL_FILE", "pendientes.xlsx")

AREAS_VALIDAS = [
    "Santiago", "Paracas", "Ingenio", "Santa Cruz",
    "Ica", "Marcona", "Subtanjalla", "Huirpacancha", "Parcona"
]

RESPONSABLES = [
    "miguel", "fernando", "pablo", "wilmer", "josselyn",
    "jaime", "gerson", "sotelo", "tiga", "william", "ochoa", "roy"
]


def inicializar_excel():
    """Crea el archivo Excel si no existe."""
    if os.path.exists(EXCEL_FILE):
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Pendientes"

    headers = ["ID", "Área", "Categoría", "Actividad", "Fecha Límite",
               "Responsable", "Estado", "Observación", "Fecha Registro"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    widths = [6, 15, 18, 50, 14, 15, 12, 35, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 9 else ""].width = w
    cols = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
    for i, col_letter in enumerate(cols):
        ws.column_dimensions[col_letter].width = widths[i]

    ws.auto_filter.ref = "A1:I1"
    ws.freeze_panes = "A2"
    wb.save(EXCEL_FILE)


def obtener_siguiente_id():
    """Obtiene el siguiente ID disponible."""
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    max_id = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] and isinstance(row[0], (int, float)):
            max_id = max(max_id, int(row[0]))
    wb.close()
    return max_id + 1


def agregar_pendiente(area, actividad, fecha_str, responsable, observacion=""):
    """Agrega un pendiente al Excel."""
    inicializar_excel()

    area_match = next((a for a in AREAS_VALIDAS if a.lower() == area.lower().strip()), None)
    if not area_match:
        sugerencias = ", ".join(AREAS_VALIDAS)
        return f"❌ Área '{area}' no válida.\nÁreas disponibles:\n{sugerencias}"

    try:
        year = datetime.now().year
        if "-" in fecha_str:
            parts = fecha_str.strip().split("-")
            fecha = datetime(year, int(parts[1]), int(parts[0]))
        elif "/" in fecha_str:
            parts = fecha_str.strip().split("/")
            fecha = datetime(year, int(parts[1]), int(parts[0]))
        else:
            fecha = None
    except (ValueError, IndexError):
        return f"❌ Fecha '{fecha_str}' no válida. Usa formato dd-mm (ej: 15-03)"

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    nuevo_id = obtener_siguiente_id()

    categoria = detectar_categoria(actividad)

    nueva_fila = [
        nuevo_id,
        area_match,
        categoria,
        actividad.strip(),
        fecha.strftime("%d/%m/%Y") if fecha else "Sin fecha",
        responsable.strip().title(),
        "Pendiente",
        observacion.strip() if observacion else "",
        datetime.now().strftime("%d/%m/%Y %H:%M")
    ]

    ws.append(nueva_fila)

    row_num = ws.max_row
    estado_cell = ws.cell(row=row_num, column=7)
    estado_cell.fill = PatternFill("solid", fgColor="FFF2CC")
    estado_cell.font = Font(bold=True)

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for col in range(1, 10):
        ws.cell(row=row_num, column=col).border = thin_border
        ws.cell(row=row_num, column=col).alignment = Alignment(vertical="center", wrap_text=True)

    wb.save(EXCEL_FILE)
    wb.close()

    fecha_txt = fecha.strftime("%d/%m/%Y") if fecha else "Sin fecha"
    return (f"✅ *Pendiente registrado*\n"
            f"📋 ID: {nuevo_id}\n"
            f"📍 Área: {area_match}\n"
            f"🏷️ Categoría: {categoria}\n"
            f"📝 Actividad: {actividad.strip()}\n"
            f"📅 Fecha: {fecha_txt}\n"
            f"👤 Responsable: {responsable.strip().title()}")


def detectar_categoria(actividad):
    """Detecta automáticamente la categoría de la actividad."""
    texto = actividad.lower()
    if any(w in texto for w in ["expediente", "ficha", "tdr"]):
        return "Expediente Técnico"
    elif any(w in texto for w in ["pago", "cobro", "valorización", "saldo", "proveedor"]):
        return "Pagos/Cobros"
    elif any(w in texto for w in ["seguimiento", "verificar", "coordinar", "revisar"]):
        return "Seguimiento"
    elif any(w in texto for w in ["presentar", "presentación", "informe"]):
        return "Presentación"
    elif any(w in texto for w in ["proceso", "selección", "bases", "propuesta"]):
        return "Proceso"
    elif any(w in texto for w in ["liquidación", "liquidacion"]):
        return "Liquidación"
    elif any(w in texto for w in ["plazo", "ampliación", "paralización", "reinicio"]):
        return "Plazos"
    elif any(w in texto for w in ["obra", "ejecución", "mano de obra", "material"]):
        return "Obra"
    else:
        return "General"


def ver_pendientes_area(area):
    """Muestra los pendientes de un área."""
    inicializar_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    area_match = next((a for a in AREAS_VALIDAS if a.lower() == area.lower().strip()), None)
    if not area_match:
        wb.close()
        return f"❌ Área '{area}' no encontrada.\nUsa /areas para ver la lista."

    pendientes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] and row[1].lower() == area_match.lower() and row[6] == "Pendiente":
            pendientes.append(row)

    wb.close()

    if not pendientes:
        return f"✅ No hay pendientes en *{area_match}*"

    msg = f"📋 *Pendientes de {area_match}* ({len(pendientes)}):\n\n"
    for p in pendientes:
        fecha = p[4] if p[4] else "Sin fecha"
        resp = p[5] if p[5] else "Sin asignar"
        msg += f"• [{p[0]}] {p[3]}\n  📅 {fecha} | 👤 {resp}\n\n"

    return msg


def ver_pendientes_hoy():
    """Muestra todos los pendientes de hoy."""
    inicializar_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    hoy = date.today().strftime("%d/%m/%Y")
    pendientes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[4] and str(row[4]).strip() == hoy and row[6] == "Pendiente":
            pendientes.append(row)

    wb.close()

    if not pendientes:
        return f"✅ No hay pendientes para hoy ({hoy})"

    msg = f"📅 *Pendientes de hoy ({hoy})*:\n\n"
    for p in pendientes:
        msg += f"• [{p[0]}] *{p[1]}*: {p[3]}\n  👤 {p[5] or 'Sin asignar'}\n\n"

    return msg


def marcar_completado(id_str):
    """Marca un pendiente como completado."""
    inicializar_excel()
    try:
        target_id = int(id_str.strip())
    except ValueError:
        return "❌ ID no válido. Usa un número (ej: /listo 5)"

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if row[0].value and int(row[0].value) == target_id:
            row[6].value = "Completado"
            row[6].fill = PatternFill("solid", fgColor="C6EFCE")
            row[6].font = Font(bold=True, color="006100")
            wb.save(EXCEL_FILE)
            wb.close()
            return f"✅ Pendiente #{target_id} marcado como *completado*"

    wb.close()
    return f"❌ No se encontró pendiente con ID {target_id}"


def generar_ayuda():
    """Genera el mensaje de ayuda."""
    return (
        "🤖 *Bot de Pendientes de Oficina*\n\n"
        "*Comandos disponibles:*\n\n"
        "📝 *Registrar actividad:*\n"
        "/nueva [área] / [actividad] / [fecha dd-mm] / [responsable]\n"
        "_Ejemplo:_\n"
        "/nueva Santiago / solicitar ampliación de plazo / 15-03 / miguel\n\n"
        "📋 *Ver pendientes por área:*\n"
        "/ver [área]\n"
        "_Ejemplo:_ /ver Paracas\n\n"
        "📅 *Ver pendientes de hoy:*\n"
        "/hoy\n\n"
        "✅ *Marcar como completado:*\n"
        "/listo [ID]\n"
        "_Ejemplo:_ /listo 5\n\n"
        "🗺️ *Ver áreas disponibles:*\n"
        "/areas\n\n"
        "❓ *Ver esta ayuda:*\n"
        "/ayuda\n\n"
        "💡 *Tip:* También puedes agregar observación:\n"
        "/nueva Santiago / tarea / 15-03 / miguel / nota importante"
    )


@app.route("/webhook", methods=["POST"])
def webhook():
    """Endpoint que recibe los mensajes de WhatsApp vía Twilio."""
    incoming_msg = request.values.get("Body", "").strip()
    resp = MessagingResponse()
    msg = resp.message()

    if not incoming_msg:
        msg.body("❌ Mensaje vacío. Envía /ayuda para ver los comandos.")
        return str(resp)

    texto = incoming_msg.lower().strip()

    if texto.startswith("/nueva"):
        contenido = incoming_msg[6:].strip()
        partes = [p.strip() for p in contenido.split("/")]
        if len(partes) < 4:
            msg.body(
                "❌ Formato incorrecto.\n\n"
                "Usa: /nueva área / actividad / fecha / responsable\n"
                "Ej: /nueva Santiago / solicitar ampliación / 15-03 / miguel"
            )
        else:
            observacion = partes[4] if len(partes) > 4 else ""
            resultado = agregar_pendiente(partes[0], partes[1], partes[2], partes[3], observacion)
            msg.body(resultado)

    elif texto.startswith("/ver"):
        area = incoming_msg[4:].strip()
        if not area:
            msg.body("❌ Indica un área. Ej: /ver Santiago\nUsa /areas para ver la lista.")
        else:
            msg.body(ver_pendientes_area(area))

    elif texto == "/hoy":
        msg.body(ver_pendientes_hoy())

    elif texto.startswith("/listo"):
        id_str = incoming_msg[6:].strip()
        if not id_str:
            msg.body("❌ Indica el ID. Ej: /listo 5")
        else:
            msg.body(marcar_completado(id_str))

    elif texto == "/areas":
        lista = "\n".join(f"• {a}" for a in AREAS_VALIDAS)
        msg.body(f"🗺️ *Áreas disponibles:*\n\n{lista}")

    elif texto in ["/ayuda", "/help", "hola", "hi", "?"]:
        msg.body(generar_ayuda())

    else:
        msg.body(
            "🤔 No entendí el comando.\n"
            "Envía /ayuda para ver las instrucciones."
        )

    return str(resp)


@app.route("/health", methods=["GET"])
def health():
    return {"status": "ok", "timestamp": datetime.now().isoformat()}


if __name__ == "__main__":
    inicializar_excel()
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
